from curses.ascii import isalpha, isdigit
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook
import requests
import html5lib

wb=Workbook()

print("************************************")
print("* GRADING SNIPPET AND INVESTIGATOR *")
print("************************************")
def funcGnc(x,sh,fn,p):
    global name,rn,an,ab
    an=[]
    ab=[]
    fname=p+"\\"+str(fn)+".xlsx"    
    sheetname=str(sh)
    sheetname = wb.create_sheet(sheetname)
    sheetname.append(['NAME','REGISTER NUMBER','20PMCA301 INTERNAL' ,'20PMCA301 EXTERNAL' ,'20PMCA301 TOTAL','20PMCA302 INTERNAL' ,'20PMCA302 EXTERNAL' ,'20PMCA302 TOTAL','20PMCA303 INTERNAL' ,'20PMCA303 EXTERNAL' ,'20PMCA303 TOTAL','20PMCA306 INTERNAL' ,'20PMCA306 EXTERNAL' ,'20PMCA306 TOTAL','20PMCA304P INTERNAL' ,'20PMCA304P EXTERNAL' ,'20PMCA304P TOTAL','20PMCA305P INTERNAL' ,'20PMCA305P EXTERNAL' ,'20PMCA305P TOTAL','19PGSL401 INTERNAL' ,'19PGSL401 EXTERNAL' ,'19PGSL401 TOTAL','20PMCA307 INTERNAL' ,'20PMCA307 EXTERNAL' ,'20PMCA307 TOTAL','20PMCA308 INTERNAL' ,'20PMCA308 EXTERNAL' ,'20PMCA308 TOTAL','20PMCA309 INTERNAL' ,'20PMCA309 EXTERNAL' ,'20PMCA309 TOTAL','20PMCA310 INTERNAL' ,'20PMCA310 EXTERNAL' ,'20PMCA310 TOTAL','20PMCA311 INTERNAL' ,'20PMCA311 EXTERNAL' ,'20PMCA311 TOTAL','20PMCA312P INTERNAL' ,'20PMCA312P EXTERNAL' ,'20PMCA312P TOTAL','20PMCA313P INTERNAL' ,'20PMCA313P EXTERNAL' ,'20PMCA313P TOTAL','19PGSL407 INTERNAL' ,'19PGSL407 EXTERNAL' ,'19PGSL407 TOTAL','20PINT401 INTERNAL' ,'20PINT401 EXTERNAL' ,'20PINT401 TOTAL','20PMCA314 INTERNAL' ,'20PMCA314 EXTERNAL' ,'20PMCA314 TOTAL','20PMCA317 INTERNAL' ,'20PMCA317 EXTERNAL' ,'20PMCA317 TOTAL','20PMCA316 INTERNAL' ,'20PMCA316 EXTERNAL' ,'20PMCA316 TOTAL','20PMCA315 INTERNAL' ,'20PMCA315 EXTERNAL' ,'20PMCA315 TOTAL','20PMCA318 INTERNAL' ,'20PMCA318 EXTERNAL' ,'20PMCA318 TOTAL','20PMCA319P INTERNAL' ,'20PMCA319P EXTERNAL' ,'20PMCA319P TOTAL','20PMCA320P INTERNAL' ,'20PMCA320P EXTERNAL' ,'20PMCA320P TOTAL','20PGSL405 INTERNAL' ,'20PGSL405 EXTERNAL' ,'20PGSL405 TOTAL','20PINT401 INTERNAL' ,'20PINT401 EXTERNAL' ,'20PINT401 TOTAL','20PMCA321 INTERNAL' ,'20PMCA321 EXTERNAL' ,'20PMCA321 TOTAL','20PMCA322 INTERNAL' ,'20PMCA322 EXTERNAL' ,'20PMCA322 TOTAL','20PMCA323 INTERNAL' ,'20PMCA323 EXTERNAL' ,'20PMCA323 TOTAL','20PGSL406 INTERNAL' ,'20PGSL406 EXTERNAL' ,'20PGSL406 TOTAL'])
    subcode=["20PMCA301","20PMCA302","20PMCA303","20PMCA306","20PMCA304P","20PMCA305P","19PGSL401","20PMCA307","20PMCA308","20PMCA309","20PMCA310","20PMCA311","20PMCA312P","20PMCA313P","19PGSL407","20PINT401","20PMCA314","20PMCA317","20PMCA316","20PMCA315","20PMCA318","20PMCA319P","20PMCA320P","20PGSL405","20PINT401","20PMCA321","20PMCA322","20PMCA323","20PGSL406"]
    ro=1
    count=0
    for xx in x:
        ro=ro+1
        count=count+1
        payload = {
        'regno': xx,
        'button': 'Get Marks'
        }
        url = 'http://results.worldcolleges.info/gncresults/nov2023/display.php'
        data = urlencode(payload)
        data = data.encode('ascii')
        req = Request(url,data)
        req.add_header('User-Agent','Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1')
        res = urlopen(req)
        lxml = BeautifulSoup(res, 'html.parser')
        scraped=lxml.find_all("tr", class_="result_table")
        res_array=[res_array.text for res_array in lxml]
        sname=res_array[3].split("\n")
        xx=(res_array[7]).split("\n")
        abc=(sname+xx)
        name=abc[3]
        rn=abc[7]
        an.append(name)
        ab.append(rn)
        sheetname.append([name,rn])
        for length in range(0,len(subcode)):
            
            if subcode[length] in abc:
                subcode_index=abc.index(subcode[length])
                subint=abc[subcode_index+1]
                subex=abc[subcode_index+2]
                subtot=abc[subcode_index+3]
                if subex =="":
                    subex="AAA"
                if subcode[length] == subcode[0]:
                        sheetname.cell(row=ro,column=3).value=subint 
                        sheetname.cell(row=ro,column=4).value=subex  
                        sheetname.cell(row=ro,column=5).value=subtot 
                if subcode[length] == subcode[1]:
                        sheetname.cell(row=ro,column=6).value=subint  
                        sheetname.cell(row=ro,column=7).value=subex   
                        sheetname.cell(row=ro,column=8).value=subtot  
                if subcode[length] == subcode[2]:
                        sheetname.cell(row=ro,column=9).value=subint  
                        sheetname.cell(row=ro,column=10).value=subex  
                        sheetname.cell(row=ro,column=11).value=subtot 
                if subcode[length] == subcode[3]:
                        sheetname.cell(row=ro,column=12).value=subint 
                        sheetname.cell(row=ro,column=13).value=subex  
                        sheetname.cell(row=ro,column=14).value=subtot
                if subcode[length] == subcode[4]:
                        sheetname.cell(row=ro,column=15).value=subint
                        sheetname.cell(row=ro,column=16).value=subex
                        sheetname.cell(row=ro,column=17).value=subtot
                if subcode[length] == subcode[5]:
                        sheetname.cell(row=ro,column=18).value=subint
                        sheetname.cell(row=ro,column=19).value=subex
                        sheetname.cell(row=ro,column=20).value=subtot
                if subcode[length] == subcode[6]:
                        sheetname.cell(row=ro,column=21).value=subint
                        sheetname.cell(row=ro,column=22).value=subex
                        sheetname.cell(row=ro,column=23).value=subtot
                if subcode[length] == subcode[7]:
                        sheetname.cell(row=ro,column=24).value=subint
                        sheetname.cell(row=ro,column=25).value=subex
                        sheetname.cell(row=ro,column=26).value=subtot
                if subcode[length] == subcode[8]:
                        sheetname.cell(row=ro,column=27).value=subint
                        sheetname.cell(row=ro,column=28).value=subex
                        sheetname.cell(row=ro,column=29).value=subtot
                if subcode[length] == subcode[9]:
                        sheetname.cell(row=ro,column=30).value=subint
                        sheetname.cell(row=ro,column=31).value=subex
                        sheetname.cell(row=ro,column=32).value=subtot
                if subcode[length] == subcode[10]:
                        sheetname.cell(row=ro,column=33).value=subint
                        sheetname.cell(row=ro,column=34).value=subex
                        sheetname.cell(row=ro,column=35).value=subtot
                if subcode[length] == subcode[11]:
                        sheetname.cell(row=ro,column=36).value=subint
                        sheetname.cell(row=ro,column=37).value=subex
                        sheetname.cell(row=ro,column=38).value=subtot
                if subcode[length] == subcode[12]:
                        sheetname.cell(row=ro,column=39).value=subint
                        sheetname.cell(row=ro,column=40).value=subex
                        sheetname.cell(row=ro,column=41).value=subtot
                if subcode[length] == subcode[13]:
                        sheetname.cell(row=ro,column=42).value=subint
                        sheetname.cell(row=ro,column=43).value=subex
                        sheetname.cell(row=ro,column=44).value=subtot
                if subcode[length] == subcode[14]:
                        sheetname.cell(row=ro,column=45).value=subint
                        sheetname.cell(row=ro,column=46).value=subex
                        sheetname.cell(row=ro,column=47).value=subtot
                if subcode[length] == subcode[15]:
                        sheetname.cell(row=ro,column=48).value=subint
                        sheetname.cell(row=ro,column=49).value=subex
                        sheetname.cell(row=ro,column=50).value=subtot
                if subcode[length] == subcode[16]:
                        sheetname.cell(row=ro,column=51).value=subint
                        sheetname.cell(row=ro,column=52).value=subex
                        sheetname.cell(row=ro,column=53).value=subtot
                if subcode[length] == subcode[17]:
                        sheetname.cell(row=ro,column=54).value=subint
                        sheetname.cell(row=ro,column=55).value=subex
                        sheetname.cell(row=ro,column=56).value=subtot
                if subcode[length] == subcode[18]:
                        sheetname.cell(row=ro,column=57).value=subint
                        sheetname.cell(row=ro,column=58).value=subex
                        sheetname.cell(row=ro,column=59).value=subtot
                if subcode[length] == subcode[19]:
                        sheetname.cell(row=ro,column=60).value=subint
                        sheetname.cell(row=ro,column=61).value=subex
                        sheetname.cell(row=ro,column=62).value=subtot
                if subcode[length] == subcode[20]:
                        sheetname.cell(row=ro,column=63).value=subint
                        sheetname.cell(row=ro,column=64).value=subex
                        sheetname.cell(row=ro,column=65).value=subtot
                if subcode[length] == subcode[21]:
                        sheetname.cell(row=ro,column=66).value=subint
                        sheetname.cell(row=ro,column=67).value=subex
                        sheetname.cell(row=ro,column=68).value=subtot
                if subcode[length] == subcode[22]:
                        sheetname.cell(row=ro,column=69).value=subint
                        sheetname.cell(row=ro,column=70).value=subex
                        sheetname.cell(row=ro,column=71).value=subtot
                if subcode[length] == subcode[23]:
                        sheetname.cell(row=ro,column=72).value=subint
                        sheetname.cell(row=ro,column=73).value=subex
                        sheetname.cell(row=ro,column=74).value=subtot
                if subcode[length] == subcode[24]:
                        sheetname.cell(row=ro,column=75).value=subint
                        sheetname.cell(row=ro,column=76).value=subex
                        sheetname.cell(row=ro,column=77).value=subtot
                if subcode[length] == subcode[25]:
                        sheetname.cell(row=ro,column=78).value=subint
                        sheetname.cell(row=ro,column=79).value=subex
                        sheetname.cell(row=ro,column=80).value=subtot
                if subcode[length] == subcode[26]:
                        sheetname.cell(row=ro,column=81).value=subint
                        sheetname.cell(row=ro,column=82).value=subex
                        sheetname.cell(row=ro,column=83).value=subtot
                if subcode[length] == subcode[27]:
                        sheetname.cell(row=ro,column=84).value=subint
                        sheetname.cell(row=ro,column=85).value=subex
                        sheetname.cell(row=ro,column=86).value=subtot

        print(str(count)+". "+name+" "+rn)
    wb.save(fname)            
   


                
               